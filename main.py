import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook("DataBase of Singer Refrigerator Model.xlsx")
worksheet = workbook["Database"]
new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

new_worksheet['A1'] = "SKU"
new_worksheet['C1'] = "MODEL"
new_worksheet['F1'] = "Date & Time"
new_worksheet['I1'] = "UNIQUE_MODELS"
new_worksheet['L1'] = "COUNT"
new_worksheet['M1'] = "TOTAL"

bold_font = Font(bold=True)
blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
for j in range(1,15):
    new_worksheet.cell(row = 1, column = j).fill = blue_fill
    new_worksheet.cell(row = 1, column = j).font = bold_font
new_worksheet.freeze_panes = new_worksheet.cell(row = 2, column = 1)
def check_cell():
    if new_worksheet.cell(row=i, column=1).value == search_value[:10]:
        return True
    else:
        return False
def print_date_time():
    now = datetime.datetime.now()
    date_string = now.strftime('%d-%m-%Y , %H:%M:%S')
    new_worksheet.cell(row=i, column=6).value = date_string
    print("DATE & TIME :", date_string)

counter = 0
previous_value = []
model_counter = {}

for i in range(2,22):
    search_value = input("SCAN HERE: ")
    print()
    if search_value in previous_value:
        print("RPEATED SCANNING ERROR")
        new_worksheet.cell(row=i, column=3, value = "REPEATED SCANNING ERROR").font = bold_font
    else:
        for row in worksheet.iter_rows():
            if row[0].value == search_value[:10]:
                # If the user input matches the value in column 1,
                # print the value of the cell in column 2 of the corresponding row
                print("MODEL:",row[1].value)
                new_worksheet.cell(row=i, column=1).value = search_value[:10]
                new_worksheet.cell(row= i, column=3).value = row[1].value

                if row[1].value in model_counter:
                    model_counter[row[1].value] += 1
                else:
                    model_counter[row[1].value] = 1

        if new_worksheet.cell(row=i, column=1).value == search_value[:10]:
                counter += 1
        else:
            print("BARCODE IS NOT FOUND IN DATABASE")
            new_worksheet.cell(row=i, column=3, value = "BARCODE IS NOT FOUND IN DATABASE").font = bold_font
            continue
        while True:
            if check_cell():
                print_date_time()
                break
        new_worksheet.cell(row=1, column=14).value = counter
        for j, model in enumerate(model_counter, start=2):
            new_worksheet.cell(row=j, column=9, value=model)
            new_worksheet.cell(row=j, column=12, value=model_counter[model])

        new_workbook.save("Inventory Storage.xlsx")
    previous_value.append(search_value)
    print()

'''
unique_models = set()
for row in new_worksheet.iter_rows(min_row = 2, min_col = 3, max_col = 3):
    model_value = row[0].value
    if model_value:
        unique_models.add(model_value)
print("Unique Models:", unique_models)


# Create a new workbook object
wb = openpyxl.Workbook()

# Select the active worksheet
ws = wb.active


# Write the output to the worksheet
my_list = [5, 6, 7, 8, 9]
for i in range(len(my_list)):
    ws.cell(row=i+1, column=1, value=my_list[i])

# Save the workbook to a new Excel file
wb.save("Inventory Storage.xlsx")
'''


'''
my_list = list(map(int, input().split()))
my_array = np.asarray()
# printing my_array
print(my_array)

for row in worksheet.iter_rows():

    cell_value = row[1].value
    print(cell_value)
'''


