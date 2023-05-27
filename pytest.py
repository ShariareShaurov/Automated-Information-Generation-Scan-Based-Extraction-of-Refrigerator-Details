import time
import openpyxl

workbook = openpyxl.load_workbook("Inventory Storage.xlsx")
worksheet = workbook.active

print()
time.sleep(10)
previous_data = []

while True:
    # Reload the workbook to get updated data
    workbook = openpyxl.load_workbook("Inventory Storage.xlsx")
    worksheet = workbook.active
    previous_data.clear()
    total = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name = row[8]
        count = row[11]
        if name is not None and count is not None:
            previous_data.append((name, count))
            total += count
    for data in previous_data:
        name, count = data
        print(name, "=", count)

    print("TOTAL: ", total)
    workbook.save("Inventory Storage.xlsx")
    print()
    time.sleep(35)

